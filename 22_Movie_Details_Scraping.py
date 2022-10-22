'''
22 - Movie Details Scraping - Excel + Selenium
- asking the new title`s IMDb link
- ( asking the row the new record should be place ) - openpyxl removes images from 2nd tab in my original excel sheet(not part of MoviePY.xlsx) -> changed to an addition excel sheet line 20 with a permanent row/cellnumber(line 32)
- collecting the movie details(title, director, stars..) from the site and adding to the excel sheet
- opening a hungarian movie site looking for the title
- if you want to test it, make sure the excel sheet links are updated (line 20, 226)
'''

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium import webdriver

from datetime import date

import sys

from openpyxl import load_workbook
wb = load_workbook('D:/Movies_New_Record.xlsx')
ws = wb.active

# BANNER
print()
k = 11
print(' Z-z-z '*k)
print()
print(' I am Db Bee! '.center(len(' Z-z-z '*k)))
print()
print(' Z-z-z '*k)
print('\n')

link = input(' Please add the new record`s IMDb link: ')
cellnumber = 3   # cellnumber = input(' Please add the row number of the new record: ') - please see the description for more information

PATH = 'C:\Program Files (x86)\chromedriver.exe'
driver = webdriver.Chrome(PATH)
driver.get(link)

# VARIABLES FOR THE READOUT - avoiding errors at the excel write stage, if the there are less than 3 records/types to add
director_1_Read = None
director_2_Read = None
director_3_Read = None
star_1_Read = None
star_2_Read = None
star_3_Read = None
genre_1_Read = None
genre_2_Read = None
genre_3_Read = None

# TAKING THE VALUES FROM THE 
# MOVIE TITLE
try:
        element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((
                By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/h1'))
        )                       
          
        titleRead = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/h1').text      
except:
        print()
        print('*** ERROR - MOVIE TITLE ***')
        print()
        driver.quit()
        sys.exit()

# YEAR OF RELEASE
try:
        yearRead = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/div/ul/li[1]/a').text # if only one item(year) there is no index in the last li[1] just li
                   
except:
        print()
        print('*** ERROR - YEAR OF RELEASE ***')
        print()

# DIRECTOR(S)
try:    
        director_1_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[3]/ul/li[1]/div/ul/li[1]/a').text

        director_2_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[3]/ul/li[1]/div/ul/li[2]/a').text

        director_3_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[3]/ul/li[1]/div/ul/li[3]/a').text   
except:
        print() # most of the time the movies have only 1 director -> would trigger an error message / not help to identify, if there is a valid error

# STAR(S)
try:    
        star_1_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[3]/ul/li[3]/div/ul/li[1]/a').text

        star_2_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[3]/ul/li[3]/div/ul/li[2]/a').text

        star_3_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[3]/ul/li[3]/div/ul/li[3]/a').text
except:
        print()
        print('*** ERROR - STARS ***') # would be triggered if the movie has less than 3 stars
        print()
        
# GENRE(S)
try:    
        genre_1_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[1]/div[1]/div[2]/a[1]/span').text

        genre_2_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[1]/div[1]/div[2]/a[2]/span').text

        genre_3_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[1]/div[1]/div[2]/a[3]/span').text
except:
        print() # would be triggered if the movie has less than 3 genres

# TAKING THE LENGTH VALUE
try:
        # taking the 2nd item(length) from "2022 1h 33m"
        movieLengthSum = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/div/ul/li[2]').text

        # if the movie has classification(pg-13): "2022 pg-13 1h 33m" taking the 3rd item
        if 'h' not in list(movieLengthSum) or 'm' not in list(movieLengthSum):
                movieLengthSum = driver.find_element(
                By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/div/ul/li[3]').text
except:
        print()
        print('*** ERROR - LENGTH ***')
        print()
       

# VALUATE AND TRANSFORM THE LENGTH VALUE(S)
lengthHour = None
lengthMinute = None
# JUST ONE ITEM LENGTH VALUE, LIKE 45m OR 2h
if len(str(movieLengthSum).split()) == 1:
        if 'h' in list(str(movieLengthSum)):
                lengthHour = str(movieLengthSum).strip('hm')    # removing the "h" or "m" values, i know, in this case, just 'h' should be there
        if 'm' in list(str(movieLengthSum)):
                lengthMinute = str(movieLengthSum).strip('hm')

# TWO ITEMS LENGTH VALUES, LIKE 2h 32m
if len(str(movieLengthSum).split()) == 2:
        lengthList = str(movieLengthSum).split()
        lengthHour = lengthList[0].strip('hm')
        lengthMinute = lengthList[1].strip('hm')


# ADDING THE VALUES TO EXCEL
# MOVIE TITLE
cell = 'C' + str(cellnumber)
ws[cell].value = titleRead
# YEAR OF RELEASE 
cellRYear = 'E' + str(cellnumber)
ws[cellRYear].value = yearRead

# DIRECTOR(S)
if director_1_Read != None:
        cellRDirector_1 = 'F' + str(cellnumber)
        ws[cellRDirector_1].value = director_1_Read

if director_2_Read != None:
        cellRDirector_2 = 'F' + str(int(cellnumber) + 1)
        ws[cellRDirector_2].value = director_2_Read

if director_3_Read != None:
        cellRDirector_3 = 'F' + str(int(cellnumber) + 1)
        ws[cellRDirector_3].value = director_3_Read

# STAR(S)
if star_1_Read != None:
        cellRStar_1 = 'G' + str(cellnumber)
        ws[cellRStar_1].value = star_1_Read

if star_2_Read != None:
        cellRStar_2 = 'G' + str(int(cellnumber) + 1)
        ws[cellRStar_2].value = star_2_Read

if star_3_Read != None:
        cellRStar_3 = 'G' + str(int(cellnumber) + 2)
        ws[cellRStar_3].value = star_3_Read

# GENRE(S)
if genre_1_Read != None:
        cellRGenre_1 = 'H' + str(cellnumber)
        ws[cellRGenre_1].value = genre_1_Read

if genre_2_Read != None:
        cellRGenre_2 = 'I' + str(cellnumber)
        ws[cellRGenre_2].value = genre_2_Read

if genre_3_Read != None:
        cellRGenre_3 = 'J' + str(cellnumber)
        ws[cellRGenre_3].value = genre_3_Read

# MOVIE LENGTH
if lengthHour != None:
        cellLengthHour = 'Q' + str(cellnumber)
        ws[cellLengthHour].value = str(lengthHour)
if lengthMinute != None:
        cellLengthMin = 'R' + str(cellnumber)
        ws[cellLengthMin].value = str(lengthMinute)

# TODAY`S DATE
today = date.today()

day = 'K' + str(cellnumber)
ws[day].value = str(today)[8:]

day = 'L' + str(cellnumber)
ws[day].value = str(today)[5:7]

day = 'M' + str(cellnumber)
ws[day].value = str(today)[0:4]

# HOW MANY TIMES SEEN FORMULA
formula = '=COUNTA(M' + str(cellnumber) + ':M' + str(int(cellnumber) + 2) + ')'  # like: =COUNTA(M6965:M6967)
day = 'N' + str(cellnumber)
ws[day].value = formula

# 1st TIME WATCHING
cellRFirst = 'O' + str(cellnumber)
ws[cellRFirst].value = '1st'        

wb.save('D:/Movies_New_Record.xlsx')

# POSTER IMAGE
try:
        poster = driver.find_element(By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[1]/div/div[1]/div/div/div[1]/img')
        posterLink = poster.get_attribute('src')
        driver.get(posterLink)
except:
        print()
        print('*** ERROR - POSTER ***')
        print()


# LOOKING FOR THE HUNGARIAN TITLE
message = ' Please press any key for the Hungarian title. '
print()
print('*' * (len(message) + 6))
print(message.center((len(message) + 6), '*'  ))
print('*' * (len(message) + 6))
print()
input()
link = 'https://www.mafab.hu/search/&search='+ ' '.join([titleRead, yearRead])
driver.get(link)

# BYE BYE BANNER
k = 6
print()
print(' Z-z-z '*k)
print()
print(' Honey added to your jar '.center(len(' Z-z-z '*k)))
print()
print(' Z-z-z '*k)
print()

sys.exit()