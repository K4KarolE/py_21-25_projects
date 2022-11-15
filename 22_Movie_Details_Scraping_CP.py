'''
22 - Movie Details Scraping - Excel + Selenium  / optimized for movies, not ideal for TV series, TV specials /
- Cross-platform: Windows, Linux 
- ask the new title`s IMDb link
- collect the movie details(title, director, stars..) from the site and add to the excel sheet
- open the poster image in the same tab (the poster image is not 'right click saveable' on IMDb by default)
- in a new browser tab look for the movie`s hungarian title
- end of process confirmation message displayed
- if you want to test it, make sure the excel sheet links are updated (around line 43, 270)
'''

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium import webdriver

from datetime import date

# from tkinter import Tk
import pyperclip as pc


# from requests import options            #headless chrome / it is slower
# options = webdriver.ChromeOptions()
# options.headless = True

import sys, webbrowser, platform

# BANNER
print()
k = 11
print(' Z-z-z '*k)
print()
print(' I aM D bee! '.center(len(' Z-z-z '*k)))
print()
print(' Z-z-z '*k)
print('\n')

link = pc.paste()
cellnumber = 3


if platform.system() == 'Windows':
        from openpyxl import load_workbook
        wb = load_workbook('D:/Movies_New_Record.xlsx')
        ws = wb.active

        PATH = 'C:\Program Files (x86)\chromedriver.exe'
        driver = webdriver.Chrome(PATH) # driver = webdriver.Chrome(PATH, chrome_options=options) #headless chrome / it is slower
        driver.minimize_window()
        driver.get(link)

if platform.system() == 'Linux':
        from openpyxl import load_workbook
        wb = load_workbook(r'/media/zsandark/D/Movies_New_Record.xlsx')
        ws = wb.active

        PATH = '/home/zsandark/Downloads/chromedriver'
        driver = webdriver.Chrome(PATH) # driver = webdriver.Chrome(PATH, chrome_options=options) #headless chrome / it is slower
        driver.minimize_window()
        driver.get(link)


# VARIABLES FOR THE READOUT - avoiding errors at the excel writing stage
# - if the there are less than 3 records/types to add
# - no length value displayed on IMDb
director_1_Read = None
director_2_Read = None
director_3_Read = None
star_1_Read = None
star_2_Read = None
star_3_Read = None
genre_1_Read = None
genre_2_Read = None
genre_3_Read = None
movieLengthSum = None

# TAKING THE VALUES FROM THE 
# MOVIE TITLE
try:
        element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((
                By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/h1'))
        )                       
          
        titleRead = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/h1').text      
except:
        print()
        print('*** ERROR - MOVIE TITLE ***')
        print()
        driver.quit()
        sys.exit()

# YEAR OF RELEASE
try:
        yearRead = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/div/ul/li[1]/a').text # if only one item(year) there is no index in the last li[1] just li
                   
except:
        print()
        print('*** ERROR - YEAR OF RELEASE ***')
        print()

# DIRECTOR(S)
try:    
        director_1_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[3]/ul/li[1]/div/ul/li[1]/a').text

        director_2_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[3]/ul/li[1]/div/ul/li[2]/a').text

        director_3_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[3]/ul/li[1]/div/ul/li[3]/a').text   
except:
        print() # most of the time the movies have only 1 director -> would trigger an error message / not help to identify, if there is a valid error

# STAR(S)
try:    
        star_1_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[3]/ul/li[3]/div/ul/li[1]/a').text

        star_2_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[3]/ul/li[3]/div/ul/li[2]/a').text

        star_3_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[3]/ul/li[3]/div/ul/li[3]/a').text
except:
        print()
        print('*** ERROR - STARS ***') # would be triggered if the movie has less than 3 stars
        print()
        
# GENRE(S)
try:    
        genre_1_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[1]/div[1]/div[2]/a[1]/span').text

        genre_2_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[1]/div[1]/div[2]/a[2]/span').text

        genre_3_Read = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/div[1]/div[1]/div[2]/a[3]/span').text
except:
        print() # would be triggered if the movie has less than 3 genres

# TAKING THE LENGTH VALUE
try:
        # taking the 2nd item(length) from "2022 1h 33m"
        movieLengthSum = driver.find_element(
        By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/div/ul/li[2]').text

        # if the movie has classification(pg-13): "2022 pg-13 1h 33m" taking the 3rd item
        if 'h' not in list(movieLengthSum) or 'm' not in list(movieLengthSum):
                movieLengthSum = driver.find_element(
                By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/div/ul/li[3]').text
except:
        print()
        print('*** ERROR - LENGTH ***')
        print()
       

# VALUATE AND TRANSFORM THE LENGTH VALUE(S)
lengthHour = None
lengthMinute = None
# JUST ONE ITEM LENGTH VALUE, LIKE 45m OR 2h
if len(str(movieLengthSum).split()) == 1:
        if 'h' in list(str(movieLengthSum)):
                lengthHour = str(movieLengthSum).strip('hm')    # removing the "h" or "m" values, i know, in this scenario, just 'h' should be fine
        if 'm' in list(str(movieLengthSum)):
                lengthMinute = str(movieLengthSum).strip('hm')

# TWO ITEMS LENGTH VALUES, LIKE 2h 32m
if len(str(movieLengthSum).split()) == 2:
        lengthList = str(movieLengthSum).split()
        lengthHour = lengthList[0].strip('hm')
        lengthMinute = lengthList[1].strip('hm')


# ADDING THE VALUES TO EXCEL
# MOVIE TITLE
cell = 'C' + str(cellnumber)
ws[cell].value = titleRead
# YEAR OF RELEASE 
cellRYear = 'E' + str(cellnumber)
ws[cellRYear].value = yearRead

# DIRECTOR(S)
cellRDirector_1 = 'F' + str(cellnumber)
ws[cellRDirector_1].value = None                # removing the previous value from the cell
if director_1_Read != None:
        ws[cellRDirector_1].value = director_1_Read

cellRDirector_2 = 'F' + str(int(cellnumber) + 1)
ws[cellRDirector_2].value = None
if director_2_Read != None:
        ws[cellRDirector_2].value = director_2_Read

cellRDirector_3 = 'F' + str(int(cellnumber) + 2)
ws[cellRDirector_3].value = None
if director_3_Read != None:
        ws[cellRDirector_3].value = director_3_Read

# STAR(S)
cellRStar_1 = 'G' + str(cellnumber)
ws[cellRStar_1].value = None                   # removing the previous value from the cell    
if star_1_Read != None:
        ws[cellRStar_1].value = star_1_Read

cellRStar_2 = 'G' + str(int(cellnumber) + 1)
ws[cellRStar_2].value = None
if star_2_Read != None:
        ws[cellRStar_2].value = star_2_Read

cellRStar_3 = 'G' + str(int(cellnumber) + 2)
ws[cellRStar_3].value = None
if star_3_Read != None:
        ws[cellRStar_3].value = star_3_Read

# GENRE(S)
cellRGenre_1 = 'H' + str(cellnumber)
ws[cellRGenre_1].value = None                   # removing the previous value from the cell
if genre_1_Read != None:
        ws[cellRGenre_1].value = genre_1_Read

cellRGenre_2 = 'I' + str(cellnumber)
ws[cellRGenre_2].value = None
if genre_2_Read != None:
        ws[cellRGenre_2].value = genre_2_Read

cellRGenre_3 = 'J' + str(cellnumber)
ws[cellRGenre_3].value = None
if genre_3_Read != None:
        ws[cellRGenre_3].value = genre_3_Read

# MOVIE LENGTH
cellLengthHour = 'Q' + str(cellnumber)
ws[cellLengthHour].value = None                 # removing the previous value from the cell
if lengthHour != None:
        ws[cellLengthHour].value = str(lengthHour)

cellLengthMin = 'R' + str(cellnumber)
ws[cellLengthMin].value = None    
if lengthMinute != None:
        ws[cellLengthMin].value = str(lengthMinute)

# TODAY`S DATE
today = date.today()

day = 'K' + str(cellnumber)
ws[day].value = str(today)[8:]

day = 'L' + str(cellnumber)
ws[day].value = str(today)[5:7]

day = 'M' + str(cellnumber)
ws[day].value = str(today)[0:4]

# HOW MANY TIMES SEEN FORMULA
formula = '=COUNTA(M' + str(cellnumber) + ':M' + str(int(cellnumber) + 2) + ')'  # like: =COUNTA(M6965:M6967)
day = 'N' + str(cellnumber)
ws[day].value = formula

# 1st TIME WATCHING
cellRFirst = 'O' + str(cellnumber)
ws[cellRFirst].value = '1st'        

# SAVE THE SHEET
openSheet = True
while openSheet == True:
        try:
                if platform.system() == 'Windows':
                        wb.save('D:/Movies_New_Record.xlsx')
                        openSheet = False
                        print('\n')

                if platform.system() == 'Linux':
                        wb.save(r'/media/zsandark/D/Movies_New_Record.xlsx')
                        openSheet = False
                        print('\n')
        except:
                print()
                input('!!! ERROR - Close your sheet and hit Enter !!!')

# POSTER IMAGE
try:
        poster = driver.find_element(By.CSS_SELECTOR, '.ipc-media--poster-l > img:nth-child(1)')
        posterLink = poster.get_attribute('src')
        webbrowser.open(posterLink)
except:
        print()
        print('*** ERROR - POSTER ***')
        print()

# LOOKING FOR THE HUNGARIAN TITLE
link = 'https://www.mafab.hu/search/&search='+ ' '.join([titleRead, yearRead])
webbrowser.open(link)

# BYE BYE BANNER
k = 6
print()
print(' Z-z-z '*k)
print()
print(' Honey added to your jar! '.center(len(' Z-z-z '*k)))
print()
print(' Z-z-z '*k)
print()
