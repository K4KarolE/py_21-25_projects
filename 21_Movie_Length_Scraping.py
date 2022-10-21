'''
21 - Movie Length Scraping - Excel + Selenium
- take an already existing title(which does not have movie length yet) + it`s release year from the MoviePY excel sheet
- look for it on IMDb
- open the first match in the roll down results = open the title`s movie database site
- take the length of the movie
- paste it into the excel sheet
- repeat it for the full sheet -> we are going to have a better estimations about the total time spent on watching movies
'''

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium import webdriver

import time, sys

from openpyxl import load_workbook
wb = load_workbook('MoviePY.xlsx')
ws = wb.active

# PLEASE WAIT BANNER
print()
k = 50
print('*'*k)
print(' Please wait, the program is loading '.center(k, '*'))
print('*'*k)

# VALUE EXTRACTION FROM THE EXCEL
cellnumber = 6      # aka row number
n = 0               # it will be an extra delay between 2 search
while True:
    try:
        PATH = 'C:\Program Files (x86)\chromedriver.exe'
        driver = webdriver.Chrome(PATH)
        driver.get('https://www.imdb.com')

        cell = 'C' + str(cellnumber)
        movietitle = ws[cell].value

        cellRYear = 'E' + str(cellnumber)
        releaseYear = ws[cellRYear].value

        cellLengthHour = 'Q' + str(cellnumber)
        movieLengthHour = ws[cellLengthHour].value

        cellLengthMin = 'R' + str(cellnumber)
        movieLengthMin = ws[cellLengthMin].value

        def isItShow():                           # movie titles including S01, S02.. or it`s release year like "1992-1999" are TV shows
            if len(str(releaseYear)) != 4:
                return True
            else:
                for i in str(movietitle).split():
                    if i[0] == 'S' and len(i) == 3 and i[1].isdigit() and i[2].isdigit():
                        return True

        # movietitle == None: empty cell(1 title in at least 3 merged cells)
        # movietitle == '-': excluding movies with no english title
        # isItShow: excluding TV shows
        # movieLengthHour != 0 and movieLengthMin != 0: look for title which does not have length yet
        while movietitle == None or movietitle == '-' or isItShow() or movieLengthHour != None or movieLengthMin != None:
            cellnumber += 1
            cell = 'C' + str(cellnumber)
            movietitle = ws[cell].value

            cellRYear = 'E' + str(cellnumber)
            releaseYear = ws[cellRYear].value

            cellLengthHour = 'Q' + str(cellnumber)
            movieLengthHour = ws[cellLengthHour].value

            cellLengthMin = 'R' + str(cellnumber)
            movieLengthMin = ws[cellLengthMin].value

        k = len(str(movietitle)) + 35  # using for error messages
        searchForMovie = ' '.join([str(movietitle), str(releaseYear)])

# IMDb SEARCH BOX
        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.ID, 'suggestion-search'))
            )
            search = driver.find_element(By.ID, 'suggestion-search')
            search.send_keys(searchForMovie)
        except:
            driver.quit()

# CLICK ON THE FIRST OF THE RESULTS
        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="react-autowhatever-1--item-0"]/a'))
            )
            search = driver.find_element(
                By.XPATH, '//*[@id="react-autowhatever-1--item-0"]/a')
            search.click()
        except:
            driver.quit()

# TAKING THE LENGTH VALUE
        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/div/ul/li[2]'))
            )

            # taking the 2nd item(length) from "2022 1h 33m"
            movieLengthSum = driver.find_element(
                By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/div/ul/li[2]').text
            
            # if the movie has classification(pg-13): "2022 pg-13 1h 33m" taking the 3rd item
            if 'h' not in list(movieLengthSum) or 'm' not in list(movieLengthSum):
                movieLengthSum = driver.find_element(
                    By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/div/ul/li[3]').text
        except:
            driver.quit()

# VALUATE AND TRANSFORM THE LENGTH VALUE(S)
        lengthHour = None
        lengthMinute = None
    # JUST ONE ITEM LENGTH VALUE, LIKE 45m OR 2h
        if len(str(movieLengthSum).split()) == 1:
            if 'h' in list(str(movieLengthSum)):
                # removing the "h" or "m" values, i know, in this case, just 'h' should ve there
                lengthHour = str(movieLengthSum).strip('hm')
            if 'm' in list(str(movieLengthSum)):
                lengthMinute = str(movieLengthSum).strip('hm')

    # TWO ITEMS LENGTH VALUES, LIKE 2h 32m
        if len(str(movieLengthSum).split()) == 2:
            lengthList = str(movieLengthSum).split()
            lengthHour = lengthList[0].strip('hm')
            lengthMinute = lengthList[1].strip('hm')

# ADDING THE VALUES TO EXCEL
        if lengthHour != None:
            ws[cellLengthHour].value = str(lengthHour)
        if lengthMinute != None:
            ws[cellLengthMin].value = str(lengthMinute)
                
        driver.close()
        cellnumber += 3  # because of the merged cells, 3 is the minimum distance between 2 titles
        time.sleep(14 + n)  # ad extra/growing delay between two search till 27 sec.
        n += 1
        if n == 14:
            n = 0

    except KeyboardInterrupt:                                     
        print()                                
        k = len(str(movietitle)) + 35
        print('*'*k)
        print((' The last movie: ' + str(movietitle) + ' (' + str(releaseYear) + ') ').center(k, '*'))
        print('*'*k)
        print()
        wb.save('MoviePY.xlsx')
        driver.quit()
        sys.exit()